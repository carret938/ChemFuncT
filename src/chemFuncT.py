"""Handles tasks associated with the Chemical Function Taxonomy sqlite database.

E. Tyler Carr
November 21, 2024
"""
from pathlib import Path
from typing import Literal

from sqlite_handler import SqliteHandler

class ChemFuncTHelper(SqliteHandler):
    """Class representing a connection to the Functional Use Classification DB."""
    
    def __init__(self, path: str | Path | None = None):
        """Constructor for ChemFuncTHelper objects.
        
        Inherits from amos.src.sqlite_handler.SqliteHandler class.
        
        Automatically instantiates a connection and cursor to the db.
        
        Parameters
        ----------
        path : str | pathlib.Path | None, default None
            The path to the ChemFuncTHelper sqlite .db file. If None, uses the parent
            class sqlite_handler.SqliteHandler attribute.
        """
        super().__init__()
        
        if path is None:
            self.set_conn(self.chem_func_uses_path)
        else:
            if not self._has_db_suffix(path):
                raise ValueError(f"path must point to a .db file, not {path}")
            path = Path(str(path)).resolve()
            self.set_conn(path)
    
    def query_hierarchy_paths(self, as_str: bool = False) -> tuple[str, str] | tuple[list[str]]:
        """Returns a string of ' -> ' delimited paths for all hierarchy paths starting
        from root nodes on. The first str contains the classification_ids, the
        second str contains the classification names.
        
        Parameters
        ----------
        as_str : bool, default True
            If True, returns results in a ' -> ' delimited string 
            (e.g., "Industrial Chemicals -> Additives -> Fillers")
            Else, returns an ordered list of the paths instead
            (e.g., ["Industrial Chemicals", "Additives", "Fillers"])
        """
        
        ## Make query
        self.cursor.execute("""
            -- Recursive call to get parent/child relationships
            WITH RECURSIVE HierarchyPaths AS (
              -- Start with the root nodes
              SELECT parent_id, child_id, child_id AS Path
              FROM ClassificationHierarchy
              WHERE parent_id is NULL
                
              UNION ALL
                
              -- For each child, append parent to the path
              SELECT h.parent_id, h.child_id, Path ||' -> '|| h.child_id
              FROM ClassificationHierarchy h
              JOIN HierarchyPaths hp ON h.parent_id = hp.child_id
            )
            -- Select the full path for leaf nodes
            SELECT Path
            FROM HierarchyPaths
            """)
        
        ## Create the ordered lists
        paths = self.cursor.fetchall()
        class_name_paths = []
        class_code_paths = []
        for path in paths:
            path = path[0]
            new_path = []
            for func_id in path.split(" -> "):
                self.cursor.execute(f"SELECT classification FROM Classifications WHERE id = ?",
                                    (func_id,))
                new_path.append(self.cursor.fetchone()[0])
            new_path = " -> ".join(new_path)
            class_name_paths.append(new_path)
            class_code_paths.append(path)
        
        ## optionally convert to str
        if as_str:
            class_name_paths = " | ".join(class_name_paths)
            class_code_paths = " | ".join(class_code_paths)
        
        return class_code_paths, class_name_paths
    
    def get_chem_name(self, dtxsid: str) -> str:
        """Returns the preferred name of dtxsid"""
        self.cursor.execute("""
            SELECT name FROM Chemicals
            WHERE dtxsid is ?;             
            """, (dtxsid,))
        
        return self.cursor.fetchone()[0]
        
    def get_class_id_from_name(self, name: str) -> str:
        """Returns the classification id that matches to name."""
        self.cursor.execute("""
               SELECT id FROM Classifications
               WHERE classification is ?;             
            """, (name,))
        
        return self.cursor.fetchone()[0]
    
    def get_class_name_from_id(self, func_id: str) -> str:
        """Returns the classification label from func_id."""
        
        self.cursor.execute("""
            SELECT classification FROM Classifications
            WHERE id is ?;             
            """, (func_id,))
        
        classification = self.cursor.fetchone()[0]
        
        return classification
        
    def get_chem_classes(self, dtxsid: str, names: bool = True, only_direct_parent: bool = False, 
                         sources: None | list[Literal["wikipedia", "appril", "drugbank", "chemexpo"]] = None, 
                         as_str: bool = True) -> str | list[str]:
        """Returns the unique classes for a given dtxsid.
        
        This method is horrendous, but it works.
        
        Parameters
        ----------
        dtxsid : str
            The DTXSID for the chemical you want the classes of.
        names : bool, default True
            If True, outputs the names of classes. Else outputs the func_ids.
        only_direct_parent : bool, default False
            If True, only returns the direct parent classes of the chemical
            (ignores hierarchically implied classes).
            Else, parents ancestry tree will be followed to get all class memberships.
        sources : None | list[literal["wikipedia", "appril", "drugbank", "chemexpo"]], default None
            A list of sources to pull classifications from. If None, all sources are used.
        as_str : bool, default True
            If True, outputs a semicolon delimited str of classes (e.g., "Drugs; Pharmaceuticals")
            Else, outputs a list of classes (e.g., ["Drugs", "Pharmaceuticals"])
        
        Returns
        -------
        str | list[str]
            The functional use classes of dtxsid (either the ids or the names).
        """
        if type(sources) == list:
            for source in sources:
                if source.lower() not in ["wikipedia", "appril", "drugbank", "chemexpo"]:
                    raise ValueError("Sources must be None or a list containing one or more of the following literals: 'wikipedia', 'appril', 'drugbank', 'chemexpo'.")
        ## get the direct parent classification_ids
        if sources == None:
            self.cursor.execute("""
                SELECT classification_id FROM ChemicalClassifications
                WHERE dtxsid = ?                
                """, (dtxsid,))
        else:
            placeholders = ', '.join('?' for _ in sources)
            self.cursor.execute(f"""
                SELECT classification_id FROM ChemicalClassifications
                WHERE dtxsid = ?         
                AND
                source_id IN ({placeholders})       
                """, (dtxsid, *sources))
        direct_parent_ids = list(set([parent_id[0] for parent_id in self.cursor.fetchall()]))
        
        ## if just looking for direct parents, we can return now
        if ((not names) and only_direct_parent):
            if as_str:
                direct_parent_ids = "; ".join(direct_parent_ids)
                
            return direct_parent_ids
        elif only_direct_parent:
            direct_parent_names = set()
            for direct_parent_id in direct_parent_ids:
                direct_parent_name = self.get_class_name_from_id(direct_parent_id)
                direct_parent_names.add(direct_parent_name)
            
            direct_parent_names = list(direct_parent_names)
            direct_parent_names.sort()
            
            if as_str:
                direct_parent_names = "; ".join(direct_parent_names)
            
            return direct_parent_names
        
        ## get the parent's parents recursively
        ancestry_ids = set()
        for parent_id in direct_parent_ids:
            self.cursor.execute("""
                WITH RECURSIVE HierarchyAncestry AS (
                    -- start ancestry with direct parent node
                    SELECT parent_id, child_id, child_id AS Ancestry
                    FROM ClassificationHierarchy
                    WHERE child_id IS ?
                    
                    UNION ALL
                    
                    -- for each ancestor, append the ancestry
                    SELECT h.parent_id, h.child_id, Ancestry ||' -> '|| h.parent_id
                    FROM ClassificationHierarchy h
                    JOIN HierarchyAncestry ha ON ((h.child_id = ha.parent_id) AND (h.parent_id IS NOT NULL))
                )
                SELECT Ancestry
                FROM HierarchyAncestry;
                """, (parent_id,))
            
            ancestry_strs = [path[0] for path in self.cursor.fetchall()]
            for path in ancestry_strs:
                for p in path.split(" -> "):
                    ancestry_ids.add(p)
                    
        ancestry_ids = list(ancestry_ids)
        ancestry_ids.sort()
        
        if not names:
            if as_str:
                ancestry_ids = "; ".join(ancestry_ids)
            return ancestry_ids
        
        ## get the names
        ancestry_names = set()
        for ancestry_id in ancestry_ids:
            ancestry_name = self.get_class_name_from_id(ancestry_id)
            ancestry_names.add(ancestry_name)
            
        ancestry_names = list(ancestry_names)
        ancestry_names.sort()
            
        if not names:
            if as_str:
                ancestry_ids = "; ".join(ancestry_ids)
                
            return ancestry_ids
        else:
            if as_str:
                ancestry_names = "; ".join(ancestry_names)
                
            return ancestry_names
    
    def get_class_parents(self, node: str, names: bool = True) -> list[str]:
        """Returns a list of parents to the node class.
        
        Parameters
        ----------
        node : str
            The name of the class of interest. Can be the class id,
            or class name -- class_id always starts with func_, so 
            it is easy to parse.
        names : bool, default True
            If True, outputs the names of classes. Else outputs the func_ids.
            
        Returns
        -------
        list[str]
            The list of parents of the functional use class node.
        """
        if node.startswith("func_"):
            self.cursor.execute("""
                SELECT parent_id FROM ClassificationHierarchy
                WHERE ((child_id is ?) AND (parent_id IS NOT NULL));               
                """, (node,))    
        else:
            self.cursor.execute("""
                SELECT parent_id FROM ClassificationHierarchy
                WHERE child_id = (
                    SELECT id FROM Classifications
                    WHERE classification = ?
                );
                """, (node,))
        
        parents = set()
        for row in self.cursor.fetchall():
            parents.add(row[0])
            
        parents = list(parents)
        
        if names:
            new_parents = set()
            for parent_id in parents:
                self.cursor.execute("""
                    SELECT classification FROM Classifications
                    WHERE id = ?;                
                    """, (parent_id,))
                new_parents.add(self.cursor.fetchone()[0])
            
            parents = list(new_parents)
            
        parents.sort()
        
        return parents
    
    def get_class_children(self, node: str, names: bool = True) -> list[str]:
        """Returns a list of children to the node class.
        
        Parameters
        ----------
        node : str
            The name of the class of interest. Can be the class id,
            or class name -- class_id always starts with func_, so 
            it is easy to parse.
        names : bool, default True
            If True, outputs the names of classes. Else outputs the func_ids.
            
        Returns
        -------
        list[str]
            The list of children of the functional use class node.
        """
        if node.startswith("func_"):
            self.cursor.execute("""
                SELECT child_id FROM ClassificationHierarchy
                WHERE ((parent_id is ?) AND (child_id IS NOT NULL));               
                """, (node,))    
        else:
            self.cursor.execute("""
                SELECT child_id FROM ClassificationHierarchy
                WHERE parent_id = (
                    SELECT id FROM Classifications
                    WHERE classification = ?
                );
                """, (node,))
        
        parents = set()
        for row in self.cursor.fetchall():
            parents.add(row[0])
            
        parents = list(parents)
        
        if names:
            new_parents = set()
            for parent_id in parents:
                self.cursor.execute("""
                    SELECT classification FROM Classifications
                    WHERE id = ?;                
                    """, (parent_id,))
                new_parents.add(self.cursor.fetchone()[0])
            
            parents = list(new_parents)
            
        parents.sort()
        
        return parents
    
    def export_db_to_excel(self, output_path: str | Path) -> None:
        """Exports the entire database to an Excel file with one worksheet per table.

        Parameters
        ----------
        output_path : str | pathlib.Path
            The path to save the resulting .xlsx file.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
        except ImportError:
            raise ImportError(
                "The 'openpyxl' library is required for this method. "
                "Please install it using 'pip install openpyxl'."
            )
        
        if not self.conn:
            raise ValueError("No active database connection. Please set a connection first.")

        ## Create a new Excel workbook
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)  # Remove the default sheet

        ## setup alignment and font objects for header formatting
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center") 

        ## Get all table names in the database
        self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = self.cursor.fetchall()

        for table in tables:
            table_name = table[0]

            ## Create a new worksheet for each table
            worksheet = workbook.create_sheet(title=table_name)

            ## Fetch all rows and column names from the table
            self.cursor.execute(f"SELECT * FROM {table_name};")
            rows = self.cursor.fetchall()
            column_names = [description[0] for description in self.cursor.description]

            ## Write column headers to the worksheet
            worksheet.append(column_names)
            
            ## Apply formatting to headers
            for column_i, column_name in enumerate(column_names, start=1):
                cell = worksheet.cell(row=1, column=column_i)
                cell.alignment = center_alignment
                cell.font = bold_font

            ## Write rows to the worksheet
            for row in rows:
                worksheet.append(row)
                
            ## Adjust column widths
            for column_i, column_name in enumerate(column_names, start=1):
                max_length = len(column_name)  # Start with the header length
                for row in rows:
                    try:
                        max_length = max(max_length, len(str(row[column_i - 1])))
                    except IndexError:
                        pass
                adjusted_width = max_length + 2  # Add some padding
                worksheet.column_dimensions[chr(64 + column_i)].width = min(adjusted_width, 110)

        ## Save the workbook to the specified output path
        workbook.save(output_path)

if __name__ == "__main__":
    pass    
